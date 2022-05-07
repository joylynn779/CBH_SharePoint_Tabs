import openpyxl
from openpyxl import load_workbook
import geopandas as gpd
import geopy
from geopy.geocoders import Nominatim
from geopy.geocoders import GoogleV3
from geopy.extra.rate_limiter import RateLimiter
import pandas as pd
import numpy as np
from geopy.geocoders import GoogleV3
import api_creds

# geolocator = GoogleV3(api_key=api_creds.google_api)


file_name = "UB_Addresses.xlsx"
tab = 'Addresses to Validate'
wrkbk = load_workbook(file_name)
sheet = wrkbk[tab]

# Example
# address = geolocator.geocode("14 N LA CIENEGA BLVD Beverly Hills CA 90211")
# print(address)
# print(address.raw)

address_list = []
# iterate through excel and display data
for i in range(2, sheet.max_row + 1):
    address_type = sheet.cell(row=i, column=1)
    address = sheet.cell(row=i, column=12)
    if address_type == 'Address':
        print(address)
        address_list.append(address)

print(address_list)
    # for j in range(1, sheet.max_column + 1):
    #     address_type = sheet.cell(row=i, column=1)
    #     if address_type='Address':
    #         cell_obj = sheet.cell(row=i, column=j)
    #         address_list.append(cell_obj)


#
# df = pd.DataFrame(['Empire State Building', 'Eiffel Tower', 'Colosseum'], columns=['Location'])
#
# print(df)
#

#         # Example
#         address = geolocator.geocode("Empire State Building")
#         address
#
#         address = normalize_address({
#             'country_code': 'US',
#             'country_area': 'California',
#             'city': 'Mountain View',
#             'postal_code': '94043',
#             'street_address': '1600 Amphitheatre Pkwy'})
#         print(address)
#
#     location_info_df = df['Location'].apply(location_info)
#     location_info_df
#
#     df_locations = pd.concat([df, location_info_df], axis=1)
#     df_locations




# # Setting up Nominatim
# locator = Nominatim(user_agent="my-application", timeout=20)
# rgeocode = RateLimiter(locator.reverse, min_delay_seconds=0.001)


# def location_info(x):
#     data = locator.geocode(x).raw
#     data_converted = pd.json_normalize(data).squeeze()  # squeeze converts a dataframe to a pandas series
#     return data_converted
#
#