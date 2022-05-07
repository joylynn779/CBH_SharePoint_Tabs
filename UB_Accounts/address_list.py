import api_creds
from geopy.geocoders import GoogleV3
from openpyxl import load_workbook
import pandas as pd


geolocator = GoogleV3(api_key=api_creds.google_api)


def location_info(x):
    data = geolocator.geocode(x).raw
    data_converted = pd.json_normalize(data).squeeze()  # squeeze converts a dataframe to a pandas series
    return data_converted


file_name = "UB_Addresses.xlsx"
tab = 'Addresses to Validate'
wrkbk = load_workbook(file_name)
sheet = wrkbk[tab]

address_list = []
# iterate through excel and display data
for i in range(2, sheet.max_row + 1):
    address_type = sheet.cell(row=i, column=1).value
    address = sheet.cell(row=i, column=12).value
    if address_type == 'Address':
        address_list.append(address)

print(address_list)

df = pd.DataFrame(address_list, columns=['Location'])

location_info_df = df['Location'].apply(location_info)

df_locations = pd.concat([df, location_info_df], axis=1)
print(df_locations)
